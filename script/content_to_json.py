import openpyxl
import json
import io
import os
import warnings
from openpyxl import Workbook
warnings.simplefilter("ignore")


# excel表格转json文件
def excel_to_json(excel_file):

    work_book = openpyxl.load_workbook(excel_file)
    # 获取workbook中全部的表格
    sheets = work_book.sheetnames
    all_articles=[]#存储全部文章
    # 循环遍历全部sheet
    for sheet_index in range(1,len(sheets)):
        json_dict = []  # json字典
        heads = []  # 表头
        sheet = work_book[sheets[sheet_index]]  # 获取excel文件中的表
        max_row = sheet.max_row
        max_column = sheet.max_column

        # 解析第一行的表头数据
        for column in range(max_column):
            heads.append(sheet.cell(1, column + 1).value)

        for row in range(max_row):  # 遍历每一行
            if row < 1:  # 跳过表头那一行
                continue

            one_line = {}  # 存储一行数据
            row_is_none = True  # 判断该行是否为空
            sentence = []  # 用于记录句子时，有多个句子的情况,仅用于第五个表格

            for column in range(max_column):  # 遍历一行中的每一个单元格
                key_from_head = heads[column]  # 获取表头做为json的key名
                if key_from_head == None:  # 跳过表头为空的列
                    continue

                cell_value = sheet.cell(row + 1, column + 1).value  # 获取该单元格的值

                if sheet_index >= 3 :  # 对句子类做特殊处理
                    if column + 1 >= 3 and cell_value != None:  # 将后面的句子追加到数组
                        line_index=cell_value.splitlines()#有换行符，则视为多个句子
                        for index in range(len(line_index)):
                            sentence.append(line_index[index])

                    if column + 1 == 3:  # 添加句子数组，只有那一栏加
                        one_line[key_from_head] = sentence
                    elif column + 1 <= 2:  # 添加前面的两个值
                        one_line[key_from_head] = cell_value

                else:
                    one_line[key_from_head] = cell_value

                if cell_value != None:  # 一行中任意一个非空即添加
                    row_is_none = False

            if row_is_none == False:
                json_dict.append(one_line)  # 添加非空的列
                if(sheet_index <= 2):#全部文章
                    all_articles.append(one_line)

        work_book.close()
        if(sheet_index==2):#生成全部文章json
            save_json_file(all_articles, '../json/全部文章.json')  # 按表名来保存文件
            print('已将表格     '+work_book.sheetnames[0] +'和'+work_book.sheetnames[1]+'     转化为全部文章.json' )
        save_json_file(json_dict, '../json/' + work_book.sheetnames[sheet_index] + ".json")  # 按表名来保存文件
        print("表格:"+work_book.sheetnames[sheet_index] + '     已成功转化到    ' + work_book.sheetnames[sheet_index] + ".json")


def save_json_file(jd, json_f_name):  # 将json保存为文件
    file = io.open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)  # 转为json
    file.write(txt)
    file.close()


def json_to_excel(json_file, excel_file):
    # 读取json文件数据
    with open(json_file, mode='r', encoding='utf-8') as jf:
        json_data = json.load(jf)
    work_book = Workbook()  # 创建excel文件
    work_sheet = work_book.active
    j = 0
    for sheet_data in json_data:
        j += 1
        i = 65  # ASCII码'65'表示'A'
        for key in sheet_data:  # 遍历要写入的行
            work_sheet['%s%d' % (chr(i), j)] = str(sheet_data[key])
            i += 1
    work_book.save(excel_file)


# 直接引用本模块时运行以下代码
if '__main__' == __name__:
    os.chdir(os.path.dirname(__file__))  # 切换工作目录到当前文件所在位置
# '''
#     json转表格
#     file_names = os.listdir('./json')  # 获取文件列表
#     for file_info in file_names:
#         print(file_info)
#         str_list = file_info.split('.')
#         print(str_list[0])
#         json_to_excel('./json/'+file_info, 'excel/'+str_list[0]+'.xlsx')
# '''
    file_names = os.listdir('../excel')  # 获取文件列表
    for file_info in file_names:  # 遍历所有文件
        # str_list = file_info.split('.')  # 以.号为结束，提取文件名
        # print(str_list)
        excel_to_json('../excel/' + file_info)
