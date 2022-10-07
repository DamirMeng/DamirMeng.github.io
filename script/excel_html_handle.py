# ------------------------------------------------------------------------------------------------
# 将excel文件内的网站数据内容转化为HTML，同时调整json文件的内容
# 在我的文章、转载文章表格中 是否新增	未处理文件名 填入值后将会按 文章链接 中名称来生成新的html文件
# 同时会在对应的内容更新到新的json到 全部文章.json 和对应的同名json中
# ------------------------------------------------------------------------------------------------
import openpyxl
import json
import io
import os
import warnings
from openpyxl import Workbook
warnings.simplefilter("ignore")

# title文章的标题
def createHTML( srcfilename,destfilename,title):  # srcfilename没处理文件
    # 三个htmlTemplate均为模板
    htmlTemplate1 = "<html lang=\"zh-CN\">"
    htmlTemplate1 += ""
    htmlTemplate1 += "<head>"
    htmlTemplate1 += "    <meta charset=\"UTF-8\">"
    htmlTemplate1 += "    <title>"

    htmlTemplate2 = "</title>"
    htmlTemplate2 += "    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">"
    htmlTemplate2 += "    <meta http-equiv=\"X-UA-Compatible\" content=\"IE=edge,chrome=1\">"
    htmlTemplate2 += "    <link href=\"../../css/base.css\" rel=\"stylesheet\">"
    htmlTemplate2 += "    <link href=\"../../css/typora.style.css\" rel=\"stylesheet\">"
    htmlTemplate2 += "    <link href=\"../../css/mobile.css\" rel=\"stylesheet\">"
    htmlTemplate2 += "    <link rel=\"icon\" sizes=\"any\" mask=\"\" href=\"https://s3.ax1x.com/2020/11/18/DneSpV.png\">"
    htmlTemplate2 += "</head>"
    htmlTemplate2 += ""
    htmlTemplate2 += "<body>"
    htmlTemplate2 += "    <!-- PC、移动端导航栏和菜单 -->"
    htmlTemplate2 += "    <script src=\"../../../../public/nav.js\"></script>"
    htmlTemplate2 += "    <div id=\"perspective\">"
    htmlTemplate2 += "        <div id=\"container\">"
    htmlTemplate2 += "            <!-- 正文，文章列表 -->"
    htmlTemplate2 += "            <main>"
    htmlTemplate2 += "                <header class=\"title\">"
    htmlTemplate2 += "                    <h2>--</h2>"
    htmlTemplate2 += "                    <div class=\"mini\"><span class=\"later\"></span><span class=\"later\">--</span>"
    htmlTemplate2 += "                        <span class=\"later\">--</span>"
    htmlTemplate2 += "                        <span class=\"hh\"><span class=\"later\">--</span>"
    htmlTemplate2 += "                            <span id=\"word\" class=\"later\">--</span></span>"
    htmlTemplate2 += "                    </div>"
    htmlTemplate2 += "                </header>"
    htmlTemplate2 += "                <div class=\"massage\">"
    htmlTemplate2 += ""
    htmlTemplate2 += ""

    htmlTemplate3 = "            </main>"
    htmlTemplate3 += "            <!-- 底部栏 -->"
    htmlTemplate3 += "            <script src=\"../../../../public/footer.js\"></script>"
    htmlTemplate3 += "        </div>"
    htmlTemplate3 += "    </div>"
    htmlTemplate3 += "    <!-- aside可置顶显示 -->"
    htmlTemplate3 += "    <script src=\"../../../../public/aside.js\"></script>"
    htmlTemplate3 += "    <script async src=\"//busuanzi.ibruce.info/busuanzi/2.3/busuanzi.pure.mini.js\"></script>"
    htmlTemplate3 += "    <script type=\"text/javascript\" src=\"../../public/base.js\"></script>"
    htmlTemplate3 += "</body>"
    htmlTemplate3 += ""
    htmlTemplate3 += "</html>"

    # 打开需要添加的文件
    contentFile = open(srcfilename, "r", encoding="utf-8")
    contentHtml = contentFile.read()
    # 组合字符串
    text = htmlTemplate1 + title+htmlTemplate2 +\
        contentHtml.split("<div class='typora-export-content'>")[1].split("</body>")[0]\
        + htmlTemplate3

    # 保存到文件
    fh = open(destfilename, 'w', encoding="utf-8")
    fh.write(text)
    fh.close()

def excel_to_json(excel_file):

    work_book = openpyxl.load_workbook(excel_file)
    # 获取workbook中全部的表格
    sheets = work_book.sheetnames
    all_articles = []  # 存储全部文章
    # 循环遍历全部sheet
    for sheet_index in range(1, len(sheets)):
        json_dict = []  # json字典
        heads = []  # 表头
        sheet = work_book[sheets[sheet_index]]  # 获取excel文件中的表
        max_row = sheet.max_row
        max_column = sheet.max_column

        # 解析第一行的表头数据到heads
        for column in range(max_column):
            heads.append(sheet.cell(1, column + 1).value)

        for row in range(max_row):  # 遍历每一行
            if row < 1:  # 跳过表头那一行
                continue

            one_line = {}  # 存储一行数据
            row_is_none = True  # 判断该行是否为空
            sentence = []  # 用于记录句子时，有多个句子的情况,仅用于第五个表格

            # 按列来遍历
            for column in range(max_column):  # 遍历一行中的每一个单元格
                key_from_head = heads[column]  # 获取表头做为json的key名
                if key_from_head == None:  # 跳过表头为空的列
                    continue

                cell_value = sheet.cell(row + 1, column + 1).value  # 获取该单元格的值

                if sheet_index >= 3:  # 对句子类做特殊处理
                    if column + 1 >= 3 and cell_value != None:  # 将后面的句子追加到数组
                        line_index = cell_value.splitlines()  # 有换行符，则视为多个句子
                        for index in range(len(line_index)):
                            sentence.append(line_index[index])

                    if column + 1 == 3:  # 添加句子数组，只有那一栏加
                        one_line[key_from_head] = sentence
                    elif column + 1 <= 2:  # 添加前面的两个值
                        one_line[key_from_head] = cell_value

                else:
                    # 添加一行数据到one_line
                    if(column < 10):
                        one_line[key_from_head] = cell_value

                    elif column == 10 and cell_value and len(cell_value)>6:  # 是新添加的,处理HTML
                        srcFile = "../markdown/" + \
                            sheet.cell(row + 1, column + 2).value + ".html"#获取源html文件
                        # 删除 是否新增	未处理文件名 两列的值
                        sheet.cell(row + 1, column + 1).value = ' '
                        sheet.cell(row + 1, column + 2).value = ' '
                        destFile = "../"+one_line[heads[8]].split("../")[-1]
                        # 创建HTML文件，参数为typora生成的HTML文件和需要生成的文件名（包含路径）
                        # one_line[heads[8]].split("../")[-1]为目标文件名和路径，不包括带../的
                        createHTML(srcFile, destFile,one_line[heads[0]])
                        print("文件:"+srcFile+"     已成功转化到  文件:", destFile)

                if cell_value != None:  # 一行中任意一个非空即添加
                    row_is_none = False

            if row_is_none == False:
                json_dict.append(one_line)  # 添加非空的列
                if(sheet_index <= 2):  # 全部文章
                    all_articles.append(one_line)
        work_book.save(excel_file)
        work_book.close()
        if(sheet_index == 2):  # 生成全部文章json
            save_json_file(all_articles, '../json/全部文章.json')  # 按表名来保存文件
            print('表格:'+work_book.sheetnames[1] + '和' +
                  work_book.sheetnames[2]+'    已成功转化到     文件:全部文章.json')
        save_json_file(json_dict, '../json/' +
                       work_book.sheetnames[sheet_index] + ".json")  # 按表名来保存文件
        print("表格:"+work_book.sheetnames[sheet_index] +
              '     已成功转化到    文件:' + work_book.sheetnames[sheet_index] + ".json")


def save_json_file(jd, json_f_name):  # 将json保存为文件
    file = io.open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)  # 转为json
    file.write(txt)
    file.close()


# 直接引用本模块时运行以下代码
if '__main__' == __name__:
    os.chdir(os.path.dirname(__file__))  # 切换工作目录到当前文件所在位置
    excel_to_json('../excel/内容表格.xlsx')
